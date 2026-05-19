# Copyright (c) 2025 Author Author
# Licensed under the Creative Commons Attribution-NonCommercial 4.0 International License (CC BY-NC 4.0)
# See the LICENSE file or https://creativecommons.org/licenses/by-nc/4.0/legalcode for details.

"""
Module : reserv.views
---------------------
Vues de l’application `reserv` :
- Réservation d’un équipement
- Visualisation du calendrier hebdomadaire
- Modification / suppression / visualisation d’une réservation
- Redirection en fonction des permissions et délais
- Accueil (vue simple)

Notes :
- L’accès est réservé aux utilisateurs authentifiés (decorator @login_required).
- Les messages utilisateur (succès/erreur) sont gérés via django.contrib.messages.
- Les états/permissions sont recalculés côté vue (ex. fenêtre de modification/suppression).
"""

# Stdlib
import json
import logging
from calendar import monthrange
from collections import defaultdict
from datetime import datetime, timedelta, date, time
from typing import Any, Dict, Iterable

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.mail import send_mail, EmailMessage
from django.db.models import Q, Sum, Count, Avg, F
from django.db.models.functions import TruncMonth, ExtractWeekDay, ExtractHour
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlencode
from django.utils.text import slugify
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST

# Third-party
try:
    from weasyprint import HTML
except ImportError:
    HTML = None

try:
    from dateutil.parser import isoparse  # facultatif si python-dateutil installé
except Exception:
    isoparse = None

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Local
from equipements.models import Equipement
from facturation.utils import decouper_reservation
from reserv.forms import ReservationForm, ReservationModificationForm
from reserv.models import Reservation
from usager.models import Affiliation, Laboratoire, Fonction, Usager, Invitation
from usager.utils import est_admin_plateforme

logger = logging.getLogger(__name__)

def _to_date(val: Any) -> date | None:
    if not val:
        return None
    # accepte "YYYY-MM-DD" ou obj Date déjà parsé
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except Exception:
        return None

def _to_bool(val: Any) -> bool | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in {"1","true","vrai","yes","oui","y"}: return True
    if s in {"0","false","faux","no","non","n"}: return False
    return None  # inconnu → None

def _to_int_list(val: Any) -> list[int]:
    if val is None or val == "":
        return []
    # si déjà liste/tuple
    if isinstance(val, (list, tuple)):
        out = []
        for x in val:
            try: out.append(int(x))
            except Exception: pass
        return out
    # si string "1,2,3" ou " 1 ; 2 "
    s = str(val).replace(";", ",")
    out = []
    for part in s.split(","):
        part = part.strip()
        if not part: continue
        try: out.append(int(part))
        except Exception: pass
    return out

def _filters_from_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalise tous les filtres attendus par _filtered_reservations().
    Ajoute au besoin d'autres clés, ici on couvre celles utilisées dans stats_query.
    """
    return {
        "date_debut": _to_date(payload.get("date_debut")),
        "date_fin":   _to_date(payload.get("date_fin")),
        "equipements": _to_int_list(payload.get("equipements")),
        # exemples possibles si tu veux étendre :
        # "assistance": _to_bool(payload.get("assistance")),
        # "statut": payload.get("statut") or None,
    }

def _filters_from_request(request) -> Dict[str, Any]:
    """
    Accepte:
      - fetch() JSON (Content-Type: application/json)
      - POST form-encoded (request.POST)
      - GET querystring (request.GET) — utile pour tests rapides
    """
    ct = request.META.get("CONTENT_TYPE", "")
    if ct.startswith("application/json"):
        try:
            raw = request.body.decode("utf-8") or "{}"
            payload = json.loads(raw)
            if not isinstance(payload, dict):
                payload = {}
        except json.JSONDecodeError:
            payload = {}
        return _filters_from_payload(payload)

    # Si ce n'est pas du JSON, on merge POST puis GET (POST prioritaire)
    data = request.POST.dict() if request.method == "POST" else request.GET.dict()

    # Supporte aussi les champs multiples "?equipements=1&equipements=2"
    # (getlist dispo en QueryDict)
    try:
        if request.method == "POST":
            eq_list = request.POST.getlist("equipements")
        else:
            eq_list = request.GET.getlist("equipements")
        if eq_list:
            data["equipements"] = eq_list
    except Exception:
        pass

    return _filters_from_payload(data)


def _parse_iso_to_local_dt(s: str):
    """
    Parse une date ISO 8601 (ex: 2025-10-29T13:30:00Z) et renvoie un datetime
    en timezone locale (aware). Retourne None si invalide.
    """
    if not s:
        return None
    try:
        if isoparse:
            dt = isoparse(s)  # gère 'Z' et offsets
        else:
            s2 = s.replace('Z', '+00:00')
            dt = datetime.fromisoformat(s2)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.utc)
        return dt.astimezone(timezone.get_current_timezone())
    except Exception:
        return None

@login_required
def reserver_equipement(request, equipement_id):
    usager = get_object_or_404(Usager, compte_utilisateur=request.user)

    # charge l’équipement
    equipement = get_object_or_404(
        Equipement.objects.only("id", "nom", "actif").prefetch_related("creneaux"),
        id=equipement_id
    )

    # tarif (si affiché dans le template)
    tarif_assistance = getattr(getattr(usager, 'affiliation', None), 'tarif_assistance', None)

    if request.method == 'POST':
        form = ReservationForm(request.POST, usager=usager, equipement=equipement, request=request)
        if form.is_valid():
            reservation = form.save(commit=False)
            reservation.usager = usager
            reservation.equipement = equipement

            # 🆕 Logique pour les réservations spéciales (Admin seulement)
            if est_admin_plateforme(request.user):
                if form.cleaned_data.get('type_reservation_maintenance'):
                    try:
                        # Maintenance: Prénom="Système", Nom="Maintenance"
                        reservation.usager = Usager.objects.get(nom="Maintenance", prenom="Système")
                    except Usager.DoesNotExist:
                        messages.warning(request, "Usager 'Système Maintenance' introuvable. Réservation attribuée à vous-même.")
                elif form.cleaned_data.get('type_reservation_enseignement'):
                    try:
                        # Enseignement: Prénom="Enseignement", Nom="Sciences Biologiques"
                        reservation.usager = Usager.objects.get(nom="Sciences Biologiques", prenom="Enseignement")
                    except Usager.DoesNotExist:
                        messages.warning(request, "Usager 'Enseignement Sciences Biologiques' introuvable. Réservation attribuée à vous-même.")

            # normalisation assistance
            if not reservation.assistance or reservation.duree_assistance_minutes is None:
                reservation.duree_assistance_minutes = 0

            try:
                reservation.full_clean()
            except ValidationError as e:
                form.add_error(None, e)
                messages.error(request, "La réservation n’a pas pu être enregistrée : vérifie dates et chevauchements.")
                return render(request, 'reserv/reserver_equipement.html', {
                    'equipement': equipement,
                    'form': form,
                    'tarif_assistance': tarif_assistance,
                    'date_retour': form.cleaned_data.get("date_debut", now().date()),
                })

            # statut + mails
            admin_emails = [admin[1] for admin in getattr(settings, 'ADMINS', [])]

            if reservation.demande_exception:
                reservation.statut = 'en_attente'
                if admin_emails:
                    admin_url = request.build_absolute_uri(
                        reverse('admin:reserv_reservation_changelist') + '?statut=en_attente'
                    )
                    send_mail(
                        subject=f"[Calendrier] Demande d'exception – {reservation.equipement.nom}",
                        message=(
                            "Une réservation avec demande d'exception a été soumise :\n\n"
                            f"Usager : {usager.prenom} {usager.nom} ({usager.courriel})\n"
                            f"Équipement : {reservation.equipement.nom}\n"
                            f"Date : {reservation.date_debut} {reservation.heure_debut} → "
                            f"{reservation.date_fin} {reservation.heure_fin}\n\n"
                            "Justification :\n"
                            f"{reservation.justification or '(aucune)'}\n\n"
                            f"{admin_url}\n"
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=admin_emails,
                        fail_silently=True,
                    )
            else:
                reservation.statut = 'a_venir'

            # mail assistance (facultatif)
            if reservation.assistance and admin_emails:
                tarif_txt = f"{tarif_assistance} $/h" if tarif_assistance is not None else "N/A"
                corps = (
                    "Une assistance a été demandée pour une réservation :

"
                    f"Usager : {usager.prenom} {usager.nom} ({usager.courriel})
"
                    f"Équipement : {reservation.equipement.nom}
"
                    f"Date : {reservation.date_debut} {reservation.heure_debut} → "
                    f"{reservation.date_fin} {reservation.heure_fin}
"
                    f"Durée assistance (min) : {reservation.duree_assistance_minutes}
"
                    f"Tarif assistance (affiliation) : {tarif_txt}
"
                )
                email_msg = EmailMessage(
                    subject=f"[Calendrier] Assistance demandée – {reservation.equipement.nom}",
                    body=corps,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=admin_emails,
                )
                try:
                    from icalendar import Calendar, Event
                    import uuid
                    cal = Calendar()
                    cal.add('prodid', '-//Core Facility//Booking System//EN')
                    cal.add('version', '2.0')
                    cal.add('method', 'REQUEST')
                    evt = Event()
                    dt_debut_assist = datetime.combine(reservation.date_debut, reservation.heure_debut)
                    dt_fin_assist = dt_debut_assist + timedelta(minutes=int(reservation.duree_assistance_minutes or 0))
                    evt.add('uid', str(uuid.uuid4()))
                    evt.add('summary', f"Assistance – {reservation.equipement.nom} ({usager.prenom} {usager.nom})")
                    evt.add('dtstart', dt_debut_assist)
                    evt.add('dtend', dt_fin_assist)
                    evt.add('description', corps)
                    cal.add_component(evt)
                    email_msg.attach('assistance.ics', cal.to_ical(), 'text/calendar')
                except ImportError:
                    pass  # icalendar not installed, email sent without attachment
                email_msg.fail_silently = True
                try:
                    email_msg.send()
                except Exception:
                    pass

            reservation.save()

            # invitations formation
            # (Géré automatiquement par le signal/save du modèle via usager.utils.creer_invitations_pour_formation)
            pass


            messages.info(request,
                          "Réservation enregistrée (validation requise)." if reservation.demande_exception
                          else "Réservation enregistrée.")
            params = urlencode({'semaine': reservation.date_debut.strftime("%Y-%m-%d")})
            url = reverse('reserv:calendrier_equipement', kwargs={'equipement_id': equipement.id})
            return redirect(f'{url}?{params}')
        else:
            messages.error(request, "Erreur dans le formulaire.")
    
    else:
        # ---------- PRÉ-REMPLISSAGE SIMPLE & FIABLE ----------
        # On NE lit que ?date_debut=YYYY-MM-DD & ?heure=HH:MM
        initial = {}
        date_str = request.GET.get('date_debut')  # ex: "2025-10-31"
        heure_str = request.GET.get('heure')      # ex: "09:00"

        if date_str:
            try:
                initial['date_debut'] = datetime.strptime(date_str, "%Y-%m-%d").date()
                initial['date_fin'] = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                pass

        if heure_str:
            try:
                h, m = map(int, heure_str.split(":"))
                initial['heure_debut_h']  = f"{h:02d}"
                initial['minute_debut_m'] = f"{(m // 10) * 10:02d}"
            except ValueError:
                pass

        form = ReservationForm(initial=initial, usager=usager, equipement=equipement, request=request)

    if not equipement.actif:
        messages.error(request, "Cet équipement est actuellement inactif. Réservation impossible.")
        return redirect('accueil')

    date_retour = now().date()
    if form.is_bound and form.is_valid():
        date_retour = form.cleaned_data.get("date_debut", date_retour)

    return render(request, 'reserv/reserver_equipement.html', {
        'equipement': equipement,
        'form': form,
        'tarif_assistance': tarif_assistance,
        'date_retour': date_retour,
    })

@login_required
def calendrier_equipement(request, equipement_id):
    """
    Vue calendrier hebdomadaire pour un équipement.

    Paramètres GET :
        - semaine (YYYY-MM-DD) : date de référence pour choisir la semaine affichée (lundi → dimanche).

    Rendu :
        - creneaux_json : créneaux autorisés (jour, heure_debut/fin)
        - reservations_json : réservations chevauchant la semaine (id, usager, début/fin ISO)
        - jours_semaine / jours_semaine_json : liste des dates de la semaine
        - heures : 0..23 (pour l’affichage de la grille)
        - variables de navigation : lundi, semaine_suivante, semaine_precedente

    Notes :
        - Les réservations affichées filtrent sur certains statuts.
        - Le calcul du lundi de la semaine suit la convention ISO (weekday() : lundi=0).
    """
    equipement = get_object_or_404(
        Equipement.objects.only("id","nom").prefetch_related("creneaux"),
        id=equipement_id
    )

    # Récupère la semaine demandée ou utilise la date d'aujourd'hui
    semaine_str = request.GET.get('semaine')
    if semaine_str:
        try:
            reference_date = datetime.strptime(semaine_str, "%Y-%m-%d").date()
        except ValueError:
            reference_date = date.today()
    else:
        reference_date = date.today()

    # Calcule le lundi de la semaine courante
    lundi = reference_date - timedelta(days=reference_date.weekday())
    jours_semaine = [lundi + timedelta(days=i) for i in range(7)]
    jours_semaine_str = [j.isoformat() for j in jours_semaine]
    dimanche = lundi + timedelta(days=6)

    # Statuts visibles dans le calendrier
    reservations = (
        Reservation.objects
        .select_related('usager', 'equipement')
        .filter(
            equipement=equipement,
            date_debut__lte=dimanche + timedelta(days=1),
            date_fin__gte=lundi,
            statut__in=["a_venir", "passee"],
        )
    )
        # .exclude(statut='annulee')  # pas nécessaire avec le filtre ci-dessus, à garder si tu préfères la ceinture+bretelles

    # 🔽 Ajoute une liste des heures de 0 à 23 pour l'affichage de la grille
    heures = list(range(24))

    # Sérialisation minimale pour le front (JS ou template)
    reservations_serialisees = [
        {
            'id': r.id,
            'usager': str(r.usager),
            'debut': r.date_debut.isoformat() + 'T' + r.heure_debut.strftime('%H:%M'),
            'fin': r.date_fin.isoformat() + 'T' + r.heure_fin.strftime('%H:%M'),
            'est_maintenance': 'maintenance' in r.usager.nom.lower() or 'maintenance' in r.usager.prenom.lower(),
            'est_enseignement': 'enseignement' in r.usager.nom.lower() or 'enseignement' in r.usager.prenom.lower(),
        }
        for r in reservations
    ]

    # Sérialisation des créneaux de l’équipement
    creneaux = equipement.creneaux.all()
    creneaux_serialises = [
        {
            'jour': c.jour,
            'heure_debut': c.heure_debut.strftime('%H:%M'),
            'heure_fin': c.heure_fin.strftime('%H:%M'),
        }
        for c in creneaux
    ]

    debug_counts = (Reservation.objects
        .filter(equipement=equipement,
                date_debut__lte=dimanche + timedelta(days=1),
                date_fin__gte=lundi)
        .values('statut')
        .annotate(n=Count('id'))
        .order_by()
    )
    logger.debug("calendrier repartition statuts: %s", list(debug_counts))

    # Récupère le tarif horaire pour l'affiliation de l'usager connecté
    from facturation.utils import get_tarif_horaire
    affiliation = request.user.usager.affiliation if hasattr(request.user, 'usager') else None
    tarif_horaire = get_tarif_horaire(equipement, affiliation)

    return render(request, 'reserv/calendrier_equipement.html', {
        'creneaux_json': creneaux_serialises,
        'equipement': equipement,
        'jours_semaine': jours_semaine,
        'jours_semaine_json': jours_semaine_str,
        'heures': heures,
        'reservations_json': reservations_serialisees,
        'lundi': lundi,
        'semaine_suivante': lundi + timedelta(days=7),
        'semaine_precedente': lundi - timedelta(days=7),
        'tarif_horaire': tarif_horaire,
    })


@login_required
def modifier_reservation(request, reservation_id):
    reservation = get_object_or_404(Reservation, id=reservation_id)
    usager = get_object_or_404(Usager, compte_utilisateur=request.user)

    est_admin = est_admin_plateforme(request.user)
    est_createur = reservation.usager == usager
    dt_now = timezone.now()
    dt_debut = timezone.make_aware(datetime.combine(reservation.date_debut, reservation.heure_debut))
    dt_limite_suppression = dt_debut + timedelta(minutes=30)

    if not est_admin and not est_createur:
        return redirect('reserv:visualiser_reservation', reservation_id=reservation.id)

    modification_possible = True
    suppression_possible = True
    if not est_admin:
        if dt_now >= dt_debut:
            modification_possible = False
            if dt_now > dt_limite_suppression:
                suppression_possible = False
            if not suppression_possible:
                messages.warning(request, "Cette réservation ne peut plus être modifiée ni supprimée.")
                return redirect('reserv:visualiser_reservation', reservation_id=reservation.id)

    if request.method == 'POST':
        form = ReservationModificationForm(
            request.POST or None,
            instance=reservation,
            usager=reservation.usager,
            equipement=reservation.equipement,
            request=request
        )
        if form.is_valid():
            modif = form.save(commit=False)

            # Valide modèle (anti-chevauchement, ordre temporel)
            try:
                modif.full_clean()
            except ValidationError as e:
                form.add_error(None, e)
                messages.error(request, "Impossible d’enregistrer : vérifie dates et chevauchements.")
                return render(request, 'reserv/modifier_reservation.html', {
                    'form': form,
                    'equipement': reservation.equipement,
                    'reservation': reservation,
                    'modification_possible': modification_possible,
                    'suppression_possible': suppression_possible,
                })

            modif.statut = 'en_attente' if modif.demande_exception else 'a_venir'
            modif.save()
            messages.success(request, "Réservation modifiée avec succès.")
            params = urlencode({'semaine': modif.date_debut.strftime("%Y-%m-%d")})
            url = reverse('reserv:calendrier_equipement', kwargs={'equipement_id': reservation.equipement.id})
            return redirect(f'{url}?{params}')
        else:
            messages.error(request, "Erreur dans le formulaire.")
    else:
        form = ReservationModificationForm(
            instance=reservation,
            usager=reservation.usager,
            equipement=reservation.equipement,
            request=request
        )

    return render(request, 'reserv/modifier_reservation.html', {
        'form': form,
        'equipement': reservation.equipement,
        'reservation': reservation,
        'modification_possible': modification_possible,
        'suppression_possible': suppression_possible,
    })


@require_POST
@login_required
def supprimer_reservation(request, reservation_id):
    reservation = get_object_or_404(Reservation, id=reservation_id)

    usager = Usager.objects.filter(compte_utilisateur=request.user).first()
    est_admin = est_admin_plateforme(request.user)
    est_createur = (usager and reservation.usager_id == usager.id)

    equip_id = reservation.equipement_id
    semaine_retour = reservation.date_debut.strftime('%Y-%m-%d')

    if not est_admin and not est_createur:
        messages.error(request, "Vous n'avez pas la permission d'annuler cette réservation.")
        return redirect(reverse('reserv:calendrier_equipement', kwargs={'equipement_id': equip_id}) + f"?semaine={semaine_retour}")

    # --- Admin = autorisation totale ---
    if not est_admin:
        dt_now = timezone.now()
        dt_debut = timezone.make_aware(datetime.combine(reservation.date_debut, reservation.heure_debut))
        if dt_now > dt_debut + timedelta(minutes=30):
            messages.error(request, "Délai d’annulation dépassé. La réservation ne peut plus être annulée.")
            return redirect(reverse('reserv:calendrier_equipement', kwargs={'equipement_id': equip_id}) + f"?semaine={semaine_retour}")

    # --- Annulation effective ---
    Reservation.objects.filter(pk=reservation.pk).exclude(statut='annulee').update(statut='annulee')
    messages.success(request, "Réservation annulée avec succès.")

    return redirect(reverse('reserv:calendrier_equipement', kwargs={'equipement_id': equip_id}) + f"?semaine={semaine_retour}")


@login_required
def visualiser_reservation(request, reservation_id):
    """
    Visualisation en lecture seule d’une réservation.
    """
    reservation = get_object_or_404(Reservation, id=reservation_id)

    return render(request, 'reserv/visualiser_reservation.html', {
        'reservation': reservation
    })


@login_required
def rediriger_reservation(request, reservation_id):
    """
    Redirige l’utilisateur soit vers la modification (si autorisé et dans les délais),
    soit vers la visualisation de la réservation.

    Règle :
        - Admin ou créateur : accès à la modification.
        - Non-admin créateur : accès à la modification tant que dt_now < (début + 30 min).
        - Sinon : visualisation seule.
    """
    reservation = get_object_or_404(Reservation, id=reservation_id)
    usager = get_object_or_404(Usager, compte_utilisateur=request.user)
    dt_now = timezone.now()
    dt_debut = timezone.make_aware(datetime.combine(reservation.date_debut, reservation.heure_debut))
    dt_limite = dt_debut + timedelta(minutes=30)

    est_admin = est_admin_plateforme(request.user)
    est_createur = reservation.usager == usager

    if est_admin or est_createur:
        # Autorisé à accéder à modifier (au moins suppression possible)
        if est_admin or dt_now < dt_limite:
            return redirect('reserv:modifier_reservation', reservation_id=reservation.id)

    # Sinon → visualisation
    return redirect('reserv:visualiser_reservation', reservation_id=reservation.id)


@login_required
def accueil(request):
    """
    Accueil de l’app `reserv`.
    On redirige simplement vers l’accueil global du projet.
    """
    return redirect('accueil')

@login_required
@user_passes_test(est_admin_plateforme)
def stats_admin(request):
    """Page stats admin – UI des filtres + conteneur résultats."""
    equipements = Equipement.objects.filter(actif=True).order_by('nom').values('id', 'nom')
    affiliations = Affiliation.objects.order_by('nom').values('id', 'nom')
    return render(request, 'reserv/stats_admin.html', {
        'equipements': equipements,
        'affiliations': affiliations,
    })

@login_required
@user_passes_test(est_admin_plateforme)
def ajax_labos(request):
    ids = request.GET.get('affiliations', '')
    aff_ids = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not aff_ids:
        return JsonResponse({'laboratoires': []})

    labs = (Laboratoire.objects
            .filter(affiliation__id__in=aff_ids)   # ⬅️ plus sûr que ...affiliation_id__in
            .order_by('nom')
            .values('id', 'nom', 'affiliation_id'))
    return JsonResponse({'laboratoires': list(labs)})

@login_required
@user_passes_test(est_admin_plateforme)
def ajax_usagers(request):
    lab_ids = [int(x) for x in request.GET.get('labos','').split(',') if x.strip().isdigit()]
    fct_ids = [int(x) for x in request.GET.get('fonctions','').split(',') if x.strip().isdigit()]
    if not lab_ids:
        return JsonResponse({'usagers': []})

    qs = Usager.objects.filter(laboratoire_id__in=lab_ids)
    if fct_ids:
        qs = qs.filter(fonction_id__in=fct_ids)

    users = (qs.order_by('nom','prenom')
               .values('id','nom','prenom','laboratoire_id','fonction_id'))
    return JsonResponse({'usagers': list(users)})

@login_required
@user_passes_test(est_admin_plateforme)
def ajax_fonctions(request):
    ids = request.GET.get('labos', '')
    labo_ids = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not labo_ids:
        return JsonResponse({'fonctions': []})

    # Fonctions réellement présentes pour les usagers des labos cochés
    fonctions = (Fonction.objects
                 .filter(usager__laboratoire_id__in=labo_ids)
                 .distinct()
                 .order_by('nom')
                 .values('id', 'nom'))
    return JsonResponse({'fonctions': list(fonctions)})

def _parse_date_or_none(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date() if s else None
    except Exception:
        return None

from openpyxl.chart import LineChart, Reference
from collections import defaultdict
from datetime import datetime, timedelta

@login_required
@user_passes_test(est_admin_plateforme)
def stats_export_equipement_xlsx(request):
    filters = _extract_filters(request)
    reservations = _filtered_reservations_for_exports(filters)
    rows = _group_aggregate(reservations, 'equipement')

    # --- période effective ---
    def _parse_date(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date() if s else None
        except Exception:
            return None

    d_start = _parse_date(filters.get('date_debut'))
    d_end   = _parse_date(filters.get('date_fin'))
    if not d_start and reservations.exists():
        d_start = min(r.date_debut for r in reservations)
    if not d_end and reservations.exists():
        d_end = max(r.date_fin for r in reservations)
    if not d_start: d_start = date.today()
    if not d_end:   d_end   = d_start

    wb = Workbook()
    ws_global = wb.active
    ws_global.title = "Résumé global"

    headers = [
        "Équipement", "Réservations", "Heures totales", "Heures usage",
        "Heures assistance", "Durée moy. (h)", "Nb assistance",
        "Formations", "Usagers distincts"
    ]
    ws_global.append(headers)
    for r in rows:
        ws_global.append([
            r['nom'], r['reservations'], r['heures_totales'],
            r['heures_usage'], r['heures_assistance'],
            r['duree_moy_heures'], r['nb_assistance'],
            r['nb_formations'],
            r['usagers_distincts']
        ])

    # ---------- Helpers ----------
    def hours_split_by_day(r):
        """Répartit la durée d'une réservation par jour (dict {date: heures})."""
        res = {}
        dt_start = datetime.combine(r.date_debut, r.heure_debut)
        dt_end   = datetime.combine(r.date_fin,   r.heure_fin)
        if dt_end <= dt_start:
            return res
        cur = dt_start
        while cur.date() <= dt_end.date():
            day_start = datetime.combine(cur.date(), datetime.min.time())
            day_end   = datetime.combine(cur.date(), datetime.max.time())
            seg_start = max(cur, day_start)
            seg_end   = min(dt_end, day_end)
            if seg_end > seg_start:
                h = (seg_end - seg_start).total_seconds() / 3600.0
                res[cur.date()] = res.get(cur.date(), 0.0) + h
            cur = day_start + timedelta(days=1)
        return res

    def add_line_chart(ws, start_row, end_row, title, anchor):
        chart = LineChart()
        chart.title = title
        chart.height = 7
        chart.width = 18
        chart.legend = None

        # Axe Y
        chart.y_axis.title = "Heures"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 24

        # Données (col B) et catégories (col A)
        data = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # Forcer style ligne non lissée + marqueurs visibles
        if chart.series:  # ✅ vérifier qu’il y a une série
            for s in chart.series:
                s.smooth = False          # pas de courbe lissée
                s.marker.symbol = "circle"  # points visibles
                s.marker.size = 4

        ws.add_chart(chart, anchor)

    # --- Onglets par équipement ---
    for r in rows:
        equip_id = r['id']
        equip_nom = r['nom'][:30]
        ws = wb.create_sheet(title=equip_nom)

        # recap
        ws.append(headers)
        ws.append([
            r['nom'], r['reservations'], r['heures_totales'],
            r['heures_usage'], r['heures_assistance'],
            r['duree_moy_heures'], r['nb_assistance'],
            r['nb_formations'],
            r['usagers_distincts']
        ])

        # tableau unique
        ws.append([])
        ws.append(["Date", "Heures d'usage"])
        hours_by_date = defaultdict(float)
        equip_resas = reservations.filter(equipement_id=equip_id)
        for resa in equip_resas:
            for d, h in hours_split_by_day(resa).items():
                if d_start <= d <= d_end:
                    hours_by_date[d] += h
        cur = d_start
        start_daily = ws.max_row + 1
        while cur <= d_end:
            ws.append([
                datetime.combine(cur, datetime.min.time()),  # 👈 vraie datetime
                round(hours_by_date.get(cur, 0.0), 2)
            ])
            ws.cell(row=ws.max_row, column=1).number_format = "yyyy-mm-dd"
            cur += timedelta(days=1)
        end_daily = ws.max_row

        # --- Graphes hebdo (col H)
        cur = d_start - timedelta(days=d_start.weekday())
        chart_index = 0
        while cur <= d_end:
            monday, sunday = cur, cur + timedelta(days=6)
            if sunday < d_start or monday > d_end:
                cur += timedelta(days=7)
                continue

            start_row = max(start_daily, start_daily + (monday - d_start).days)
            end_row   = min(end_daily, start_row + 6)
            if end_row < start_row:
                cur += timedelta(days=7)
                continue

            # chaque graphe est décalé de 15 lignes
            anchor_row = 5 + chart_index * 15
            anchor = f"H{anchor_row}"
            add_line_chart(ws, start_row, end_row,
                           f"Usage hebdo – {monday.isoformat()}",
                           anchor)
            chart_index += 1
            cur += timedelta(days=7)

        # --- Graphes mensuels (col U)
        y, m = d_start.year, d_start.month
        chart_index = 0
        while date(y, m, 1) <= d_end:
            first_day = date(y, m, 1)
            last_day = date(y, m, monthrange(y, m)[1])
            if last_day < d_start or first_day > d_end:
                if m == 12: y, m = y+1, 1
                else: m += 1
                continue

            start_row = max(start_daily, start_daily + (first_day - d_start).days)
            end_row   = min(end_daily, start_row + (last_day - first_day).days)
            if end_row < start_row:
                if m == 12: y, m = y+1, 1
                else: m += 1
                continue

            # chaque graphe est décalé de 20 lignes
            anchor_row = 5 + chart_index * 20
            anchor = f"U{anchor_row}"
            add_line_chart(ws, start_row, end_row,
                           f"Usage mensuel – {y}-{m:02d}",
                           anchor)
            chart_index += 1
            if m == 12: y, m = y+1, 1
            else: m += 1

    # réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="stats_equipements.xlsx"'
    wb.save(response)
    return response
    
    
@login_required
@user_passes_test(est_admin_plateforme)
def stats_export_dimension_xlsx(request):
    filters = _extract_filters(request)
    reservations = _filtered_reservations_for_exports(filters)
    level = _last_selected_level(filters) or 'equipement'
    rows = _group_aggregate(reservations, level)

    wb = Workbook()
    ws = wb.active
    ws.title = "Résumé global"

    headers = [
        "Nom", "Réservations", "Heures totales", "Heures usage",
        "Heures assistance", "Durée moy. (h)", "Nb assistance",
        "Formations", "Usagers distincts"
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r['nom'], r['reservations'], r['heures_totales'],
            r['heures_usage'], r['heures_assistance'],
            r['duree_moy_heures'], r['nb_assistance'],
            r['nb_formations'],
            r['usagers_distincts']
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="stats_dimension.xlsx"'
    wb.save(response)
    return response



def _filters_from_request_export(request):
     def _to_ids(key):
         raw = request.POST.get(key) or request.GET.get(key) or ''
         return [int(x) for x in raw.split(',') if x.strip().isdigit()]

     return {
         'date_debut': _parse_date_or_none(request.POST.get('date_debut') or request.GET.get('date_debut')),
         'date_fin':   _parse_date_or_none(request.POST.get('date_fin')   or request.GET.get('date_fin')),
         'equipements':   _to_ids('equipements'),
         'affiliations':  _to_ids('affiliations'),
         'laboratoires':  _to_ids('laboratoires'),
         'fonctions':     _to_ids('fonctions'),
         'usagers':       _to_ids('usagers'),
     }

def _filtered_reservations(filters):
    """
    Sélectionne les réservations qui chevauchent la période (si fournie),
    filtrées par les cases cochées.
    Inclut les réservations physiquement terminées (date/heure fin passée)
    même si le cron n'a pas encore mis 'passee'.
    """
    _today = date.today()
    _now_time = datetime.now().time()
    qs = (Reservation.objects
          .select_related('usager__laboratoire__affiliation', 'usager__fonction', 'equipement')
          .exclude(statut='annulee')
          .filter(
              Q(statut='passee')
              | Q(date_fin__lt=_today)
              | Q(date_fin=_today, heure_fin__lte=_now_time)
          ))

    debut = filters['date_debut']
    fin   = filters['date_fin']
    if debut and fin:
        # chevauchement : debut <= r.date_fin && fin >= r.date_debut
        qs = qs.filter(Q(date_debut__lte=fin) & Q(date_fin__gte=debut))
    elif debut:
        qs = qs.filter(date_fin__gte=debut)
    elif fin:
        qs = qs.filter(date_debut__lte=fin)

    if filters.get('equipements'):
        qs = qs.filter(equipement_id__in=filters['equipements'])
    if filters.get('affiliations'):
        qs = qs.filter(usager__laboratoire__affiliation_id__in=filters['affiliations'])
    if filters.get('laboratoires'):
        qs = qs.filter(usager__laboratoire_id__in=filters['laboratoires'])
    if filters.get('fonctions'):
        qs = qs.filter(usager__fonction_id__in=filters['fonctions'])
    if filters.get('usagers'):
        qs = qs.filter(usager_id__in=filters['usagers'])

    return qs

def _minutes_totales(resa):
    dt1 = datetime.combine(resa.date_debut, resa.heure_debut)
    dt2 = datetime.combine(resa.date_fin, resa.heure_fin)
    return max(int((dt2 - dt1).total_seconds() // 60), 0)

def _minutes_usage_assistance(resa):
    """Retourne (min_usage, min_assistance) selon règles métier."""
    total = _minutes_totales(resa)
    if getattr(resa, 'est_formation', False):
        return (0, 0)  # forfait global → pas d'heures comptées
    # standard
    min_ass = int(getattr(resa, 'duree_assistance_minutes', 0) or 0) if getattr(resa, 'assistance', False) else 0
    return (total, min_ass)

def _agg_metrics(reservations):
    n = 0
    min_usage = 0
    min_ass   = 0
    users = set()
    nb_ass = 0
    _form_pks = []
    for r in reservations:
        n += 1
        u,a = _minutes_usage_assistance(r)
        min_usage += u
        min_ass   += a
        if r.usager_id:
            users.add(r.usager_id)
        if r.assistance:
            nb_ass += 1
        if r.est_formation:
            _form_pks.append(r.pk)

    nb_form = Invitation.objects.filter(
        reservation_id__in=_form_pks,
        date_validation__isnull=False,
    ).count() if _form_pks else 0

    total_min = min_usage + min_ass
    return {
        'reservations': n,
        'heures_totales': round(total_min/60.0, 2),
        'heures_usage': round(min_usage/60.0, 2),
        'heures_assistance': round(min_ass/60.0, 2),
        'duree_moy_heures': round((total_min/n)/60.0, 2) if n else 0.0,
        'nb_assistance': nb_ass,
        'nb_formations': nb_form,
        'usagers_distincts': len(users),
    }

def _group_key_and_label(resa, level):
    """
    Retourne (id, label) selon le niveau demandé: 'equipement' | 'affiliation' | 'laboratoire' | 'fonction' | 'usager'.
    """
    if level == 'equipement' and resa.equipement:
        return (resa.equipement_id, resa.equipement.nom)
    if level == 'affiliation' and getattr(resa.usager, 'laboratoire', None) and resa.usager.laboratoire.affiliation:
        aff = resa.usager.laboratoire.affiliation
        return (aff.id, aff.nom)
    if level == 'laboratoire' and getattr(resa.usager, 'laboratoire', None):
        lab = resa.usager.laboratoire
        return (lab.id, lab.nom)
    if level == 'fonction' and getattr(resa.usager, 'fonction', None):
        f = resa.usager.fonction
        return (f.id, f.nom)
    if level == 'usager' and resa.usager:
        u = resa.usager
        return (u.id, f"{u.nom} {u.prenom}")
    return (None, None)

def _group_aggregate(reservations, level):
    """
    Agrège par 'level' et retourne une liste de lignes: [{'id':..,'nom':.., metrics...}, ...]
    """
    from collections import defaultdict
    buckets = defaultdict(list)
    for r in reservations:
        gid, label = _group_key_and_label(r, level)
        if gid is None:
            continue
        buckets[(gid, label)].append(r)

    rows = []
    for (gid, label), group in buckets.items():
        m = _agg_metrics(group)
        rows.append({
            'id': gid,
            'nom': label,
            **m
        })
    # tri par heures totales desc puis nb réservations
    rows.sort(key=lambda x: (-x['heures_totales'], -x['reservations']))
    return rows

def _last_selected_level(filters):
    """
    Détermine le “dernier niveau coché” côté filtre.
    Priorité : usagers > fonctions > laboratoires > affiliations ; sinon None (global).
    """
    if filters['usagers']:
        return 'usager'
    if filters['fonctions']:
        return 'fonction'
    if filters['laboratoires']:
        return 'laboratoire'
    if filters['affiliations']:
        return 'affiliation'
    return None

# --- relit les filtres depuis GET (exports) ---
def _extract_filters(request):
    """Unifie les filtres pour GET et POST."""
    def _ids(name):
        raw = request.GET.get(name) or request.POST.get(name) or ''
        return [int(x) for x in raw.split(',') if x.strip().isdigit()]
    return {
        'date_debut': request.GET.get('date_debut') or request.POST.get('date_debut') or '',
        'date_fin': request.GET.get('date_fin') or request.POST.get('date_fin') or '',
        'equipements': _ids('equipements'),
        'affiliations': _ids('affiliations'),
        'laboratoires': _ids('laboratoires'),
        'fonctions': _ids('fonctions'),
        'usagers': _ids('usagers'),
    }


# --- convertit filtres (str) en dates et applique sur Reservation (statut=passee par défaut) ---
def _filtered_reservations_for_exports(filters):
    qs = (
        Reservation.objects
        .select_related('usager__laboratoire__affiliation', 'usager__fonction', 'equipement')
        .filter(statut='passee')
    )
    
    d1 = filters.get('date_debut') or ''
    d2 = filters.get('date_fin') or ''
    debut = fin = None
    if d1:
        try:    debut = datetime.strptime(d1, "%Y-%m-%d").date()
        except: pass
    if d2:
        try:    fin = datetime.strptime(d2, "%Y-%m-%d").date()
        except: pass

    if debut and fin:
        qs = qs.filter(Q(date_debut__lte=fin) & Q(date_fin__gte=debut))
    elif debut:
        qs = qs.filter(date_fin__gte=debut)
    elif fin:
        qs = qs.filter(date_debut__lte=fin)

    if filters['equipements']:
        qs = qs.filter(equipement_id__in=filters['equipements'])
    if filters['affiliations']:
        qs = qs.filter(usager__laboratoire__affiliation_id__in=filters['affiliations'])
    if filters['laboratoires']:
        qs = qs.filter(usager__laboratoire_id__in=filters['laboratoires'])
    if filters['fonctions']:
        qs = qs.filter(usager__fonction_id__in=filters['fonctions'])
    if filters['usagers']:
        qs = qs.filter(usager_id__in=filters['usagers'])

    return qs

# --- agrégation générique par dimension ---
def _rows_by_dimension(qs, dimension):
    """
    dimension ∈ {"equipement","affiliation","laboratoire","fonction","usager"}
    Return: list[dict] avec clés:
      nom, reservations, heures_usage, heures_assistance, heures_totales,
      duree_moy_heures, nb_assistance, usagers_distincts
    """
    from collections import defaultdict
    buckets = defaultdict(lambda: {
        'reservations': 0,
        'minutes_usage': 0,
        'minutes_ass': 0,
        'nb_assistance': 0,
        'usagers': set(),
    })

    def _key_name(r):
        if dimension == 'equipement':
            return (getattr(r.equipement, 'id', None), getattr(r.equipement, 'nom', '—'))
        if dimension == 'affiliation':
            aff = getattr(getattr(getattr(r, 'usager', None), 'laboratoire', None), 'affiliation', None)
            return (getattr(aff, 'id', None), getattr(aff, 'nom', '—'))
        if dimension == 'laboratoire':
            lab = getattr(getattr(r, 'usager', None), 'laboratoire', None)
            return (getattr(lab, 'id', None), getattr(lab, 'nom', '—'))
        if dimension == 'fonction':
            f = getattr(getattr(r, 'usager', None), 'fonction', None)
            return (getattr(f, 'id', None), getattr(f, 'nom', '—'))
        if dimension == 'usager':
            u = getattr(r, 'usager', None)
            nom = f"{getattr(u, 'nom', '')} {getattr(u, 'prenom', '')}".strip() or '—'
            return (getattr(u, 'id', None), nom)
        return (None, '—')

    for r in qs:
        k_id, k_nom = _key_name(r)
        usage, ass = _minutes_usage_assistance(r)
        b = buckets[(k_id, k_nom)]
        b['reservations'] += 1
        b['minutes_usage'] += usage
        b['minutes_ass'] += ass
        if getattr(r, 'assistance', False):
            b['nb_assistance'] += 1
        if r.usager_id:
            b['usagers'].add(r.usager_id)

    rows = []
    for (k_id, nom), b in buckets.items():
        total_min = b['minutes_usage'] + b['minutes_ass']
        rows.append({
            'nom': nom,
            'reservations': b['reservations'],
            'heures_usage': round(b['minutes_usage'] / 60.0, 2),
            'heures_assistance': round(b['minutes_ass'] / 60.0, 2),
            'heures_totales': round(total_min / 60.0, 2),
            'duree_moy_heures': round((total_min / b['reservations']) / 60.0, 2) if b['reservations'] else 0,
            'nb_assistance': b['nb_assistance'],
            'usagers_distincts': len(b['usagers']),
        })
    # tri : heures totales desc, puis réservations desc
    rows.sort(key=lambda x: (-x['heures_totales'], -x['reservations'], x['nom']))
    return rows

# --- EXPORT PRINCIPAL ---

def _equipement_rows(qs):
    """Ici, on agrège par équipement uniquement (une ligne par équipement)."""
    return _rows_by_dimension(qs, 'equipement')



@login_required
@user_passes_test(est_admin_plateforme)
def stats_query(request):
    filters = _filters_from_request(request)

    qs = _filtered_reservations(filters)
    resas = list(qs)

    # --- Période normalisée ---
    dd = filters['date_debut']
    df = filters['date_fin']
    if not dd and resas:
        dd = min(r.date_debut for r in resas)
    if not df and resas:
        df = max(r.date_fin for r in resas)

    def _weekday_counts(d1, d2):
        counts = {i: 0 for i in range(7)}
        if not d1 or not d2 or d1 > d2:
            return counts
        cur = d1
        while cur <= d2:
            counts[cur.weekday()] += 1
            cur += timedelta(days=1)
        return counts

    wd_counts = _weekday_counts(dd, df)

    # Helpers durée
    def _combine_safe(d, h):
        # évite les crashes si heure manquante
        return datetime.combine(d, h or time(0, 0))

    def minutes_resa(r):
        debut = _combine_safe(r.date_debut, r.heure_debut)
        fin   = _combine_safe(r.date_fin,   r.heure_fin)
        return max(int((fin - debut).total_seconds() // 60), 0)

    def split_minutes_usage_assistance(r):
        """Logique alignée sur la facturation."""
        duree = minutes_resa(r)
        # bugfix: 'resa' → 'r'
        if getattr(r, 'est_formation', False):
            return (0, 0)   # (et tu 'continue' plus bas, donc c’est redondant mais safe)
        assist_min = int(getattr(r, 'duree_assistance_minutes', 0) or 0) if getattr(r, 'assistance', False) else 0
        assist_min = max(assist_min, 0)
        return duree, assist_min

    # KPI globaux
    total_reservations = len(resas)
    total_minutes_usage = 0
    total_minutes_assist = 0
    nb_assistance = 0
    usagers_distincts = set()
    demandes_exception = 0

    # Groupes
    par_eq  = defaultdict(lambda: {'reservations': 0, 'min_usage': 0, 'min_assist': 0})
    par_labo= defaultdict(lambda: {'reservations': 0, 'min_usage': 0, 'min_assist': 0})
    par_usr = defaultdict(lambda: {'reservations': 0, 'min_usage': 0, 'min_assist': 0})

    weekly = defaultdict(lambda: {'reservations': 0, 'minutes': 0})

    for r in resas:
        mu, ma = split_minutes_usage_assistance(r)
        total_minutes_usage  += mu
        total_minutes_assist += ma

        if getattr(r, 'assistance', False):
            nb_assistance += 1

        if getattr(r, 'est_formation', False):
            continue

        if getattr(r, 'demande_exception', False):
            demandes_exception += 1

        if r.usager_id:
            usagers_distincts.add(r.usager_id)

        # Groupes
        if r.equipement:
            key = (r.equipement_id, r.equipement.nom)
            par_eq[key]['reservations'] += 1
            par_eq[key]['min_usage']    += mu
            par_eq[key]['min_assist']   += ma

        if getattr(r, 'usager', None) and getattr(r.usager, 'laboratoire', None):
            key = (r.usager.laboratoire.id, r.usager.laboratoire.nom)
            par_labo[key]['reservations'] += 1
            par_labo[key]['min_usage']    += mu
            par_labo[key]['min_assist']   += ma

        if getattr(r, 'usager', None):
            key = (r.usager.id, f"{r.usager.nom} {r.usager.prenom}")
            par_usr[key]['reservations'] += 1
            par_usr[key]['min_usage']    += mu
            par_usr[key]['min_assist']   += ma

        # Hebdo (lundi ISO)
        d0 = r.date_debut
        monday = d0 - timedelta(days=d0.weekday())
        wk = monday.isoformat()
        weekly[wk]['reservations'] += 1
        weekly[wk]['minutes']      += (mu + ma)  # total “activité” (usage + assistance)

    def _tabify(d):
        items = []
        for (id_, nom), v in d.items():
            tot = v['min_usage'] + v['min_assist']
            avg = (tot / v['reservations']) if v['reservations'] else 0
            items.append({
                'id': id_,
                'nom': nom,
                'reservations': v['reservations'],
                'heures_usage': round(v['min_usage'] / 60.0, 2),
                'heures_assistance': round(v['min_assist'] / 60.0, 2),
                'heures_totales': round(tot / 60.0, 2),
                'duree_moy_heures': round(avg / 60.0, 2),
            })
        items.sort(key=lambda x: (-x['heures_totales'], -x['reservations'], x['nom']))
        return items

    tables = {
        'equipements':  _tabify(par_eq),
        'laboratoires': _tabify(par_labo),
        'usagers':      _tabify(par_usr),
    }

    ts_weekly = [{'semaine': k, 'reservations': v['reservations'], 'heures': round(v['minutes']/60.0, 2)}
                 for k, v in sorted(weekly.items(), key=lambda kv: kv[0])]

    from usager.models import Invitation
    _form_resa_ids = [r.pk for r in resas if r.est_formation]
    nb_form = Invitation.objects.filter(
        reservation_id__in=_form_resa_ids,
        date_validation__isnull=False,
    ).count()

    kpis = {
        'reservations': total_reservations,
        'heures_totales': round((total_minutes_usage + total_minutes_assist)/60.0, 2),
        'heures_usage': round(total_minutes_usage/60.0, 2),
        'heures_assistance': round(total_minutes_assist/60.0, 2),
        'nb_assistance': nb_assistance,
        'nb_formations': nb_form,
        'usagers_distincts': len(usagers_distincts),
        'demandes_exception': demandes_exception,
        'periode': {
            'debut': dd.isoformat() if dd else '',
            'fin':   df.isoformat() if df else '',
        },
    }

    return JsonResponse({
        'ok': True,
        'kpis': kpis,
        'tables': tables,
        'timeseries': {'weekly': ts_weekly},
    })

# Palette discrète (10 couleurs): on re-map par id d'équipement

@login_required
@user_passes_test(est_admin_plateforme)
def stats_zone1(request):
    """
    Statistiques globales plateforme (zone 1) :
    - Nombre d usagers actifs
    - Nombre de laboratoires representes (au moins un usager actif)
    - Nombre de nouveaux inscrits sur la periode (reglement accepte)
    """
    from usager.models import Usager, Laboratoire

    def _parse(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date() if s else None
        except Exception:
            return None

    debut = _parse(request.GET.get("date_debut") or "")
    fin   = _parse(request.GET.get("date_fin") or "")

    nb_actifs = Usager.objects.filter(est_actif=True).count()
    nb_labos  = Laboratoire.objects.filter(usager__est_actif=True).distinct().count()

    qs_inscrits = Usager.objects.filter(reglement_accepte=True)
    if debut:
        qs_inscrits = qs_inscrits.filter(reglement_accepte_at__date__gte=debut)
    if fin:
        qs_inscrits = qs_inscrits.filter(reglement_accepte_at__date__lte=fin)
    nb_inscrits = qs_inscrits.count()

    return JsonResponse({
        "usagers_actifs":    nb_actifs,
        "labos_representes": nb_labos,
        "nouveaux_inscrits": nb_inscrits,
    })

PALETTE = [
    "#1D4ED8", "#059669", "#DC2626", "#7C3AED", "#EA580C",
    "#0E7490", "#16A34A", "#B45309", "#C026D3", "#4B5563",
]

def _equip_color(equipement_id: int) -> str:
    return PALETTE[equipement_id % len(PALETTE)]

@user_passes_test(est_admin_plateforme)
def calendrier_global_admin(request):
    """
    Page HTML : calendrier global (FullCalendar).
    """
    # Légende (équipement -> couleur)
    equips = Equipement.objects.only('id', 'nom').order_by('nom')
    legend = [{"id": e.id, "nom": e.nom, "color": _equip_color(e.id)} for e in equips]
    return render(request, "admin/calendrier_global_admin.html", {"legend": legend})

@user_passes_test(est_admin_plateforme)
def calendrier_global_admin_data(request):
    """
    Endpoint JSON pour FullCalendar.
    Attend des params GET optionnels: start, end (ISO8601) et filtres.
    """
    # Fenêtre demandée par FullCalendar (UTC ISO) -> on prend des dates locales
    start = request.GET.get("start")  # ex: 2025-10-01T00:00:00Z
    end = request.GET.get("end")
    try:
        # on reste en date (localdate) pour coller à ton projet existant
        start_date = timezone.localdate() if not start else datetime.fromisoformat(start.replace("Z","")).date()
        end_date = start_date if not end else datetime.fromisoformat(end.replace("Z","")).date()
    except Exception:
        start_date = timezone.localdate()
        end_date = start_date

    qs = (
        Reservation.objects
        .select_related('usager', 'equipement')
        .filter(
            date_debut__lte=end_date,
            date_fin__gte=start_date,
        )
        .exclude(statut__in=['annulee'])  # on cache les annulées
        .order_by('date_debut', 'heure_debut')
    )

    events = []
    for r in qs:
        # start / end ISO local sans TZ (cohérent avec tes autres vues)
        start_iso = f"{r.date_debut}T{r.heure_debut}"
        end_iso   = f"{r.date_fin}T{r.heure_fin}"
        color = _equip_color(r.equipement_id)

        title = f"{r.usager.nom_complet if hasattr(r.usager,'nom_complet') else r.usager} – {r.equipement.nom}"
        # Bordure selon statut pour différencier rapidement
        borderColor = "#16A34A" if r.statut == "a_venir" else ("#EF4444" if r.statut == "passee" else "#F59E0B")

        events.append({
            "id": r.id,
            "title": title,
            "start": start_iso,
            "end": end_iso,
            "color": color,
            "borderColor": borderColor,
            "extendedProps": {
                "equipement": r.equipement.nom,
                "usager": str(r.usager),
                "statut": r.statut,
            },
        })

    return JsonResponse(events, safe=False)


from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import openpyxl.styles.colors

from django.db.models import Sum, Count, F
from django.db.models.functions import TruncMonth

from collections import defaultdict

@login_required
@user_passes_test(est_admin_plateforme)
def stats_export_unified_xlsx(request):
    """
    Génère un rapport Excel unifié contenant :
    1. Dashboard (KPIs + Graphiques)
    2. Données Brutes (Liste complète des réservations)
    """
    # 1. Récupération des données
    filters = _extract_filters(request)
    
    # On reconstruit la query pour éviter le .only() de _filtered_reservations_for_exports
    # qui bloque le select_related sur usager__laboratoire
    qs = Reservation.objects.filter(statut='passee')
    
    d1 = filters.get('date_debut') or ''
    d2 = filters.get('date_fin') or ''
    debut = fin = None
    if d1:
        try:    debut = datetime.strptime(d1, "%Y-%m-%d").date()
        except: pass
    if d2:
        try:    fin = datetime.strptime(d2, "%Y-%m-%d").date()
        except: pass

    if debut and fin:
        qs = qs.filter(Q(date_debut__lte=fin) & Q(date_fin__gte=debut))
    elif debut:
        qs = qs.filter(date_fin__gte=debut)
    elif fin:
        qs = qs.filter(date_debut__lte=fin)

    if filters['equipements']:
        qs = qs.filter(equipement_id__in=filters['equipements'])
    if filters['affiliations']:
        qs = qs.filter(usager__laboratoire__affiliation_id__in=filters['affiliations'])
    if filters['laboratoires']:
        qs = qs.filter(usager__laboratoire_id__in=filters['laboratoires'])
    if filters['fonctions']:
        qs = qs.filter(usager__fonction_id__in=filters['fonctions'])
    if filters['usagers']:
        qs = qs.filter(usager_id__in=filters['usagers'])

    reservations = qs.select_related(
        'usager', 'usager__laboratoire', 'usager__laboratoire__affiliation', 'equipement'
    ).order_by('date_debut')

    # 2. Création du classeur
    wb = Workbook()
    
    # --- Onglet 2 : Données Brutes (On le crée en premier pour l'avoir dispo, mais on le déplacera après) ---
    ws_data = wb.active
    ws_data.title = "Données Brutes"
    
    headers = [
        "ID", "Date Début", "Date Fin", "Équipement", 
        "Usager", "Laboratoire", "Affiliation", 
        "Type", "Durée (h)", "Coût ($)"
    ]
    ws_data.append(headers)
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ============================================================================
    # PHASE 1: Agrégation des données pour KPIs et graphiques
    # ============================================================================
    from facturation.utils import decouper_reservation
    
    # Dictionnaires agrégés
    data_by_affiliation = {}  # affiliation_nom -> total_heures
    data_by_labo = {}         # labo_nom -> total_heures
    data_by_usager = {}       # usager_nom -> total_heures
    data_by_type = {'Formation': 0.0, 'Assistance': 0.0, 'Usage': 0.0}
    data_by_nature = {'Recherche': 0.0, 'Maintenance': 0.0, 'Enseignement': 0.0}

    
    # Métriques globales
    total_heures = 0.0
    total_heures_assistance = 0.0
    _form_pks_dash = []
    nb_assistances = 0
    
    # Heatmap: (jour_idx, heure) -> set(num_semaine)
    semaines_occupees = {}
    
    # Calculer le nombre de semaines dans la période
    debut = filters.get('date_debut')
    fin = filters.get('date_fin')
    if debut and fin:
        try:
            d_debut = datetime.strptime(debut, "%Y-%m-%d").date()
            d_fin = datetime.strptime(fin, "%Y-%m-%d").date()
            nb_semaines_totales = max(1, (d_fin - d_debut).days // 7)
        except:
            d_debut = None
            d_fin = None
            nb_semaines_totales = 1
    else:
        d_debut = None
        d_fin = None
        nb_semaines_totales = 1
    
    # Stocker les données pour le remplissage ultérieur
    reservation_data = []
    
    # Parcourir toutes les réservations pour agréger les données
    for r in reservations:
        # Calcul durée et coût via facturation
        segment = decouper_reservation(r)
        duree_h = float(segment.get("usage_heures", 0)) + float(segment.get("assistance_heures", 0))
        heures_assist = float(segment.get("assistance_heures", 0))
        cout = float(segment.get("usage_cout", 0)) + float(segment.get("assistance_cout", 0)) + float(segment.get("formation_cout", 0))
        
        # Métriques globales
        total_heures += duree_h
        total_heures_assistance += heures_assist
        
        # Type de réservation (exclusif)
        if r.est_formation:
            type_resa = "Formation"
            data_by_type['Formation'] += duree_h
            _form_pks_dash.append(r.pk)
        elif r.assistance:
            type_resa = "Assistance"
            data_by_type['Assistance'] += duree_h
            nb_assistances += 1
        else:
            type_resa = "Usage"
            data_by_type['Usage'] += duree_h
        
        # Affiliation
        aff_nom = "Inconnu"
        if r.usager and r.usager.laboratoire and r.usager.laboratoire.affiliation:
            aff_nom = r.usager.laboratoire.affiliation.nom
        data_by_affiliation[aff_nom] = data_by_affiliation.get(aff_nom, 0.0) + duree_h
        
        # Laboratoire
        labo_nom = "Inconnu"
        if r.usager and r.usager.laboratoire:
            labo_nom = r.usager.laboratoire.nom
        data_by_labo[labo_nom] = data_by_labo.get(labo_nom, 0.0) + duree_h
        
        if r.usager:
            usager_nom = f"{r.usager.prenom} {r.usager.nom}"
        data_by_usager[usager_nom] = data_by_usager.get(usager_nom, 0.0) + duree_h

        # Nature (Standard / Maintenance / Enseignement)
        # Basé sur l'identité de l'usager (cf. PROJECT_CONTEXT.md)
        if r.usager and r.usager.prenom == "Système" and r.usager.nom == "Maintenance":
            data_by_nature['Maintenance'] += duree_h
        elif r.usager and r.usager.prenom == "Enseignement" and r.usager.nom == "Sciences Biologiques":
            data_by_nature['Enseignement'] += duree_h
        else:
            data_by_nature['Recherche'] += duree_h
        
        
        # Heatmap: marquer TOUTES les heures occupées par cette réservation
        if d_debut and d_fin:
            # Itérer sur toutes les heures de la réservation
            dt_debut = datetime.combine(r.date_debut, r.heure_debut)
            dt_fin = datetime.combine(r.date_fin, r.heure_fin)
            
            current_dt = dt_debut
            while current_dt < dt_fin:
                jour_idx = current_dt.weekday()
                heure = current_dt.hour
                date_courante = current_dt.date()
                num_semaine = (date_courante - d_debut).days // 7
                
                key = (jour_idx, heure)
                if key not in semaines_occupees:
                    semaines_occupees[key] = set()
                semaines_occupees[key].add(num_semaine)
                
                current_dt += timedelta(hours=1)
        
        # Stocker pour remplissage ultérieur
        reservation_data.append({
            'id': r.id,
            'date_debut': r.date_debut,
            'date_fin': r.date_fin,
            'equipement': r.equipement.nom,
            'usager': usager_nom,
            'laboratoire': labo_nom,
            'affiliation': aff_nom,
            'type': type_resa,
            'duree_h': duree_h,
            'cout': cout
        })
    
    nb_formations = Invitation.objects.filter(
        reservation_id__in=_form_pks_dash,
        date_validation__isnull=False,
    ).count() if _form_pks_dash else 0

    # ============================================================================
    # PHASE 2: Remplissage de l'onglet "Données Brutes"
    # ============================================================================
    for data in reservation_data:
        row = [
            data['id'],
            data['date_debut'],
            data['date_fin'],
            data['equipement'],
            data['usager'],
            data['laboratoire'],
            data['affiliation'],
            data['type'],
            round(data['duree_h'], 2),
            round(data['cout'], 2)
        ]
        ws_data.append(row)

    # Ajustement largeur colonnes
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_data.column_dimensions[column].width = adjusted_width

    # --- Onglet 1 : Dashboard ---
    ws_dash = wb.create_sheet("Tableau de Bord", 0)
    ws_dash.sheet_view.showGridLines = False
    
    # Titre
    title_cell = ws_dash.cell(row=1, column=1, value="Rapport d'Activité - Synthèse")
    title_cell.font = Font(size=18, bold=True, color="2F5496")
    
    # Période
    p_debut = filters.get('date_debut') or "Début"
    p_fin = filters.get('date_fin') or "Fin"
    ws_dash.cell(row=2, column=1, value=f"Période : {p_debut} au {p_fin}").font = Font(italic=True, size=12)

    # --- KPIs ---
    last_row = ws_data.max_row
    if last_row > 1:
        # Ligne 1 de KPIs
        ws_dash.cell(row=4, column=1, value="Réservations").font = Font(bold=True)
        ws_dash.cell(row=5, column=1, value=f"=COUNTA('Données Brutes'!A2:A{last_row})").font = Font(size=14, color="2F5496")
        
        ws_dash.cell(row=4, column=2, value="Heures Totales").font = Font(bold=True)
        ws_dash.cell(row=5, column=2, value=f"=SUM('Données Brutes'!I2:I{last_row})").font = Font(size=14, color="2F5496")
        ws_dash.cell(row=5, column=2).number_format = '0.00'

        ws_dash.cell(row=4, column=3, value="Coût Total ($)").font = Font(bold=True)
        ws_dash.cell(row=5, column=3, value=f"=SUM('Données Brutes'!J2:J{last_row})").font = Font(size=14, color="2F5496")
        ws_dash.cell(row=5, column=3).number_format = '#,##0.00 $'
        
        ws_dash.cell(row=4, column=4, value="Usagers Distincts").font = Font(bold=True)
        nb_usagers = reservations.values('usager').distinct().count()
        ws_dash.cell(row=5, column=4, value=nb_usagers).font = Font(size=14, color="2F5496")
        
        # Ligne 2 de KPIs (NOUVEAUX)
        ws_dash.cell(row=7, column=1, value="Durée Moy/Résa (h)").font = Font(bold=True)
        duree_moyenne = total_heures / len(reservation_data) if len(reservation_data) > 0 else 0
        ws_dash.cell(row=8, column=1, value=round(duree_moyenne, 2)).font = Font(size=14, color="2F5496")
        ws_dash.cell(row=8, column=1).number_format = '0.00'
        
        ws_dash.cell(row=7, column=2, value="Formations").font = Font(bold=True)
        ws_dash.cell(row=8, column=2, value=nb_formations).font = Font(size=14, color="2F5496")
        
        ws_dash.cell(row=7, column=3, value="Assistances").font = Font(bold=True)
        ws_dash.cell(row=8, column=3, value=nb_assistances).font = Font(size=14, color="2F5496")
        
        ws_dash.cell(row=7, column=4, value="% H Assistance").font = Font(bold=True)
        pct_assistance = (total_heures_assistance / total_heures * 100) if total_heures > 0 else 0
        ws_dash.cell(row=8, column=4, value=round(pct_assistance, 1)).font = Font(size=14, color="2F5496")
        ws_dash.cell(row=8, column=4).number_format = '0.0"%"'
    else:
        ws_dash.cell(row=5, column=1, value=0)

    # --- Top 3 Heures de Pointe ---
    # Calculer le taux d'occupation pour chaque case
    heatmap_taux = {}
    for (jour_idx, heure), semaines in semaines_occupees.items():
        taux = len(semaines) / nb_semaines_totales
        heatmap_taux[(jour_idx, heure)] = taux
    
    # Trier et prendre les 3 meilleures
    top_slots = sorted(
        [(d, h, taux) for (d, h), taux in heatmap_taux.items()],
        key=lambda x: -x[2]
    )[:3]
    
    # Formater et afficher
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    if top_slots:
        top_text = " | ".join([
            f"{jours[d]} {h}h-{h+1}h ({taux*100:.0f}%)"
            for d, h, taux in top_slots
        ])
        ws_dash.cell(row=3, column=1, value=f"Heures de pointe : {top_text}").font = Font(italic=True, size=11)

    # --- Heatmap (Jours x Heures) - Taux d'Occupation ---
    ws_dash.cell(row=55, column=6, value="Heatmap - Taux d'Occupation Moyen").font = Font(bold=True)
    
    hours = list(range(8, 21)) # 8h à 20h
    days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    
    # En-têtes Heures
    for i, h in enumerate(hours):
        c = ws_dash.cell(row=56, column=7+i, value=f"{h}h")
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
    
    # En-têtes Jours
    for i, d in enumerate(days):
        ws_dash.cell(row=57+i, column=6, value=d).font = Font(bold=True)

    # Remplissage avec taux d'occupation
    start_row = 57
    start_col = 7
    
    for d_idx in range(7):
        for h_idx, h in enumerate(hours):
            key = (d_idx, h)
            taux = heatmap_taux.get(key, 0.0)
            cell = ws_dash.cell(row=start_row+d_idx, column=start_col+h_idx, value=round(taux, 2))
            if taux == 0:
                cell.font = Font(color="D3D3D3") # Griser les 0

    # Conditional Formatting (Bleu -> Orange -> Rouge)
    end_col_letter = get_column_letter(start_col + len(hours) - 1)
    range_ref = f"{get_column_letter(start_col)}{start_row}:{end_col_letter}{start_row+6}"
    
    from openpyxl.formatting.rule import ColorScaleRule
    rule = ColorScaleRule(start_type='min', start_color='4472C4',  # Bleu
                          mid_type='percentile', mid_value=50, mid_color='FFC000',  # Orange
                          end_type='max', end_color='C00000')  # Rouge
    ws_dash.conditional_formatting.add(range_ref, rule)

    # --- Données sources (Déplacées en dessous de la Heatmap pour navigation verticale) ---
    # La Heatmap finit vers la ligne 63. Le graphique d'évolution est en A64 (taille ~15 lignes).
    # On place les données à partir de la ligne 85.
    data_start_row = 85
    
    # --- Graphique 1 : Répartition par Affiliation (Camembert) ---
    col_aff = 1 # A
    ws_dash.cell(row=data_start_row, column=col_aff, value="Affiliation")
    ws_dash.cell(row=data_start_row, column=col_aff+1, value="Heures")
    
    # Utiliser les données agrégées
    row_idx = data_start_row + 1
    for aff_nom, heures in sorted(data_by_affiliation.items(), key=lambda x: -x[1]):
        ws_dash.cell(row=row_idx, column=col_aff, value=aff_nom)
        c = ws_dash.cell(row=row_idx, column=col_aff+1, value=heures)
        c.number_format = '#,##0.00"h"'  # Format personnalisé pour afficher "h"
        row_idx += 1
    
    if row_idx > data_start_row + 1:
        pie = PieChart()
        pie.title = "Répartition par Affiliation (Heures)"
        pie.height = 10
        pie.width = 18
        labels = Reference(ws_dash, min_col=col_aff, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_aff+1, min_row=data_start_row, max_row=row_idx-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # Étiquettes : Valeur + Pourcentage uniquement (pas de nom de série ni catégorie)
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.separator = ", "
        ws_dash.add_chart(pie, "A10")

    # --- Graphique 2 : Top 5 Laboratoires + Autres (Camembert) ---
    col_lab = 4 # D
    ws_dash.cell(row=data_start_row, column=col_lab, value="Laboratoire")
    ws_dash.cell(row=data_start_row, column=col_lab+1, value="Heures")
    
    # Trier et prendre top 5
    sorted_labos = sorted(data_by_labo.items(), key=lambda x: -x[1])
    top5_labos = sorted_labos[:5]
    autres_heures = sum(h for _, h in sorted_labos[5:])
    
    row_idx = data_start_row + 1
    for labo_nom, heures in top5_labos:
        ws_dash.cell(row=row_idx, column=col_lab, value=labo_nom)
        c = ws_dash.cell(row=row_idx, column=col_lab+1, value=heures)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if autres_heures > 0:
        ws_dash.cell(row=row_idx, column=col_lab, value="Autres")
        c = ws_dash.cell(row=row_idx, column=col_lab+1, value=autres_heures)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if row_idx > data_start_row + 1:
        pie = PieChart()
        pie.title = "Top 5 Laboratoires (Heures)"
        pie.height = 10
        pie.width = 18
        labels = Reference(ws_dash, min_col=col_lab, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_lab+1, min_row=data_start_row, max_row=row_idx-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # Étiquettes : Valeur + Pourcentage uniquement
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.separator = ", "
        ws_dash.add_chart(pie, "J10")

    # --- Graphique 3 : Top 10 Usagers + Autres (Camembert) ---
    col_user = 7 # G
    ws_dash.cell(row=data_start_row, column=col_user, value="Usager")
    ws_dash.cell(row=data_start_row, column=col_user+1, value="Heures")
    
    # Trier et prendre top 10
    sorted_usagers = sorted(data_by_usager.items(), key=lambda x: -x[1])
    top10_usagers = sorted_usagers[:10]
    autres_heures_usagers = sum(h for _, h in sorted_usagers[10:])
    
    row_idx = data_start_row + 1
    for usager_nom, heures in top10_usagers:
        ws_dash.cell(row=row_idx, column=col_user, value=usager_nom)
        c = ws_dash.cell(row=row_idx, column=col_user+1, value=heures)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if autres_heures_usagers > 0:
        ws_dash.cell(row=row_idx, column=col_user, value="Autres")
        c = ws_dash.cell(row=row_idx, column=col_user+1, value=autres_heures_usagers)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if row_idx > data_start_row + 1:
        pie = PieChart()
        pie.title = "Top 10 Usagers (Heures)"
        pie.height = 10
        pie.width = 18
        labels = Reference(ws_dash, min_col=col_user, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_user+1, min_row=data_start_row, max_row=row_idx-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # Étiquettes : Valeur + Pourcentage uniquement
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.separator = ", "
        ws_dash.add_chart(pie, "A32")

    # --- Graphique 4 : Type de Réservation (Camembert) ---
    col_type = 10 # J
    ws_dash.cell(row=data_start_row, column=col_type, value="Type")
    ws_dash.cell(row=data_start_row, column=col_type+1, value="Heures")
    
    # Utiliser les données agrégées
    row_idx = data_start_row + 1
    for type_nom, heures in sorted(data_by_type.items(), key=lambda x: -x[1]):
        ws_dash.cell(row=row_idx, column=col_type, value=type_nom)
        c = ws_dash.cell(row=row_idx, column=col_type+1, value=heures)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if row_idx > data_start_row + 1:
        pie = PieChart()
        pie.title = "Répartition par Type (Heures)"
        pie.height = 10
        pie.width = 18
        labels = Reference(ws_dash, min_col=col_type, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_type+1, min_row=data_start_row, max_row=row_idx-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # Étiquettes : Valeur + Pourcentage uniquement
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.separator = ", "
        pie.dataLabels.separator = ", "
        ws_dash.add_chart(pie, "J32")

    # --- Graphique 5 : Nature de l'Usage (Camembert) [NOUVEAU] ---
    col_nature = 16 # P
    ws_dash.cell(row=data_start_row, column=col_nature, value="Nature")
    ws_dash.cell(row=data_start_row, column=col_nature+1, value="Heures")
    
    # Utiliser les données agrégées
    row_idx = data_start_row + 1
    for nat_nom, heures in sorted(data_by_nature.items(), key=lambda x: -x[1]):
        ws_dash.cell(row=row_idx, column=col_nature, value=nat_nom)
        c = ws_dash.cell(row=row_idx, column=col_nature+1, value=heures)
        c.number_format = '#,##0.00"h"'
        row_idx += 1
    
    if row_idx > data_start_row + 1:
        pie = PieChart()
        pie.title = "Nature de l'Usage (Heures)"
        pie.height = 10
        pie.width = 18
        labels = Reference(ws_dash, min_col=col_nature, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_nature+1, min_row=data_start_row, max_row=row_idx-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        # Étiquettes : Valeur + Pourcentage uniquement
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        pie.dataLabels.showCatName = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.separator = ", "
        ws_dash.add_chart(pie, "S10")  # Positionné en haut à droite

    # --- Graphique 6 : Évolution Temporelle (Ligne) ---
    # Déterminer la granularité selon la période
    if d_debut and d_fin:
        periode_jours = (d_fin - d_debut).days
    else:
        periode_jours = 30  # Par défaut
    
    col_evo = 13 # M
    
    if periode_jours < 31:
        # Par jour
        ws_dash.cell(row=data_start_row, column=col_evo, value="Jour")
        ws_dash.cell(row=data_start_row, column=col_evo+1, value="Heures")
        
        # Agréger par jour
        data_by_day = {}
        for data in reservation_data:
            jour = data['date_debut']
            data_by_day[jour] = data_by_day.get(jour, 0.0) + data['duree_h']
        
        row_idx = data_start_row + 1
        for jour in sorted(data_by_day.keys()):
            ws_dash.cell(row=row_idx, column=col_evo, value=jour.strftime('%Y-%m-%d'))
            ws_dash.cell(row=row_idx, column=col_evo+1, value=round(data_by_day[jour], 2))
            row_idx += 1
        
        chart_title = "Évolution Journalière (Heures)"
    elif periode_jours < 365:
        # Par mois
        ws_dash.cell(row=data_start_row, column=col_evo, value="Mois")
        ws_dash.cell(row=data_start_row, column=col_evo+1, value="Heures")
        
        # Agréger par mois
        data_by_month = {}
        for data in reservation_data:
            mois = data['date_debut'].replace(day=1)
            data_by_month[mois] = data_by_month.get(mois, 0.0) + data['duree_h']
        
        row_idx = data_start_row + 1
        for mois in sorted(data_by_month.keys()):
            ws_dash.cell(row=row_idx, column=col_evo, value=mois.strftime('%Y-%m'))
            ws_dash.cell(row=row_idx, column=col_evo+1, value=round(data_by_month[mois], 2))
            row_idx += 1
        
        chart_title = "Évolution Mensuelle (Heures)"
    else:
        # Par trimestre
        ws_dash.cell(row=data_start_row, column=col_evo, value="Trimestre")
        ws_dash.cell(row=data_start_row, column=col_evo+1, value="Heures")
        
        # Agréger par trimestre
        data_by_quarter = {}
        for data in reservation_data:
            quarter = (data['date_debut'].year, (data['date_debut'].month - 1) // 3 + 1)
            data_by_quarter[quarter] = data_by_quarter.get(quarter, 0.0) + data['duree_h']
        
        row_idx = data_start_row + 1
        for (year, q) in sorted(data_by_quarter.keys()):
            ws_dash.cell(row=row_idx, column=col_evo, value=f"{year}-Q{q}")
            ws_dash.cell(row=row_idx, column=col_evo+1, value=round(data_by_quarter[(year, q)], 2))
            row_idx += 1
        
        chart_title = "Évolution Trimestrielle (Heures)"
    
    if row_idx > data_start_row + 1:
        # Utiliser BarChart pour les données journalières, LineChart pour mensuel/trimestriel
        if periode_jours < 31:
            chart = BarChart()
            chart.type = "col"  # Barres verticales
            # Configurer l'axe Y
            chart.y_axis.title = "Heures"
            chart.y_axis.delete = False
            chart.y_axis.majorTickMark = "out"
            chart.y_axis.tickLblPos = "nextTo"
        else:
            chart = LineChart()
            chart.style = 12
            chart.y_axis.title = "Heures"
        
        chart.title = chart_title
        
        cats = Reference(ws_dash, min_col=col_evo, min_row=data_start_row+1, max_row=row_idx-1)
        data = Reference(ws_dash, min_col=col_evo+1, min_row=data_start_row, max_row=row_idx-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_dash.add_chart(chart, "A64")

    # 3. Réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Rapport_Activite_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
